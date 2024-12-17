import numpy
import SimpleITK as sitk
import six

from radiomics import firstorder, getTestCase, glcm, glrlm, glszm, imageoperations, shape, gldm

from PyQt5.QtCore import pyqtSignal, pyqtSlot, Qt, QThread, QObject
import numpy as np
from MACRO import *
import os
import nibabel as nib
from openpyxl import Workbook
from openpyxl.styles import Font
import cv2


class TextureAnalysis(QObject):
    analysis_of_patch_finished_sig = pyqtSignal(dict) # sends the measured vals in a dict
    analysis_of_slice_finished_sig = pyqtSignal(int) # sends the slice number
    analysis_of_wholeScan_started_sig = pyqtSignal(dict) # sends a dict that includes: scan file name, scan number in dataset
    analysis_of_wholeScan_finished_sig = pyqtSignal(dict) # sends a dict that includes: scan file name, scan number in dataset
    analysis_of_dataset_finished_sig = pyqtSignal(dict)

    def __init__(self):
        super().__init__()
        self.image : sitk.Image
        self.mask : sitk.Image
        self._run_flag = 1
        self.settings = {'binWidth': 25,
                    'interpolator': sitk.sitkBSpline,
                    'resampledPixelSpacing': None}
        # self.setFeatureCalculationSettings()
        self.latest_analysis_results_dict = {'Slice':0, 'Row':0, 'Col':0, 'mean':0}
        self.create_results_dict_template({'available_texture_analysis_methods': ['GLCM', 'GLRLM', 'GLDM', 'GLSZM', 'FIRST_ORDER'], 
                                           'radiomics_algs_settings':{'binWidth': 25, 'interpolator': sitk.sitkBSpline,'resampledPixelSpacing': None},
                                           '':''})

        

    @pyqtSlot(np.ndarray)
    def set_single_image_and_mask(self, params_struct):       
        self.image_prcessor.filters_struct = params_struct
        print ('change val to: ', self.image_prcessor.filters_struct.exposure.param_dict["VAL"])

    @pyqtSlot(np.ndarray, np.ndarray, dict)
    def start_single_image_analysis_slot(self, image, mask, setting):       
        print ('start single image analysis')
        self.latest_analysis_results_dict = self.run_texture_analysis_algs(image, mask, setting).copy()
        print('Done!')
        self.analysis_of_patch_finished_sig.emit(self.latest_analysis_results_dict)

    @pyqtSlot(str, dict)
    def start_datset_image_analysis_slot(self, dataset_path, analysis_setting_dict : dict):       
        
        file_index = 0
        number_of_available_files = self.check_number_of_scanfiles_in_path(dataset_path)

        if not os.path.exists(analysis_setting_dict['save_path']):
            os.mkdir(analysis_setting_dict['save_path'])

        for scan_folder in os.listdir(dataset_path):
            scan_folder_path = os.path.join(dataset_path, scan_folder)
            for file_name in os.listdir(scan_folder_path):
           
                if 'img' in file_name and (file_name.endswith(".nii.gz") or file_name.endswith(".nii")) :
                    file_path = os.path.join(scan_folder_path, file_name)
                    mask_path = os.path.join(scan_folder_path, file_name.replace('img_', 'liver_'))
                    file_name = file_name.split('.')[0]

                    self.current_3dimage = nib.load(file_path).get_fdata()
                    self.current_3dmask =  nib.load(mask_path).get_fdata()
                    result_excel_file = analysis_setting_dict['save_path'] + 'TextureAnalysisResults_' + file_name +'.xlsx'
                    wb = Workbook()         
                    wb.remove(wb.active)
                    ws1 = wb.create_sheet('Info')
                    ws1['A1'] = 'Scan File Name'
                    ws1['B1'] = file_name
                    ws1['A2'] = 'Mask File Name'
                    # ws1['B2'] = 
                    
                    sheet_analysis_results_z_axis = wb.create_sheet('Z_Axis_AnalysisResults')
                    sheet_analysis_results_z_axis.append(list(self.results_dict_template.keys())) # slice i row j column k

                    wb.save(result_excel_file)
                    self.analysis_of_wholeScan_started_sig.emit({"file_name":file_name, "scan_number_in_dataset":file_index, "total_scans":number_of_available_files})
                    params_list = list(self.results_dict_template.keys())

                    ## run on the image:
                    if np.max(self.current_3dmask) > 0:
                        self.latest_analysis_results_dict = self.run_texture_analysis_algs(self.current_3dimage, self.current_3dmask, analysis_setting_dict).copy()
                    else: #checking for empty slices
                        self.latest_analysis_results_dict = self.results_dict_template.copy()
                    self.latest_analysis_results_dict['patch_location'] = f"whole_3d_image"
                    print(f"whole_3d_image{file_name} done")
                    current_row = sheet_analysis_results_z_axis.max_row+1
                    for param in self.latest_analysis_results_dict.keys():
                        if param in params_list:
                            sheet_analysis_results_z_axis.cell(row=current_row, column=params_list.index(param)+1,
                                                            value= str(self.latest_analysis_results_dict[param]) )
                    wb.save(result_excel_file)
                    ###################

                    for i in range(self.current_3dimage.shape[2]):
                        self.current_slice = self.current_3dimage[:, :, i]
                        self.currnet_mask_slice = self.current_3dmask[:, :, i]
                        ## run on the mask:
                        if np.max(self.currnet_mask_slice) > 0:
                            self.latest_analysis_results_dict = self.run_texture_analysis_algs(self.current_slice, self.currnet_mask_slice, analysis_setting_dict).copy()
                        else: #checking for empty slices
                            self.latest_analysis_results_dict = self.results_dict_template.copy()
                        self.latest_analysis_results_dict['patch_location'] = f"roi_slice{i}"
                        print(f"roi_slice{i} done")
                        current_row = sheet_analysis_results_z_axis.max_row+1
                        for param in self.latest_analysis_results_dict.keys():
                            if param in params_list:
                                sheet_analysis_results_z_axis.cell(row=current_row, column=params_list.index(param)+1,
                                                                value= str(self.latest_analysis_results_dict[param]) )
                        wb.save(result_excel_file)
                        ###################

                        for j in range(analysis_setting_dict['num_rows']):
                            for k in range(analysis_setting_dict['num_cols']):  
                                img_height, img_width = self.current_3dimage.shape[0], self.current_3dimage.shape[1]
                                patch_height, patch_width = img_height // analysis_setting_dict['num_rows'], img_width // analysis_setting_dict['num_cols']
                                start_point_x = int (k * patch_width)
                                start_point_y = int (j * patch_height)
                                end_point_x = int ( (k+1) * patch_width)
                                end_point_y = int ( (j+1) * patch_height)
                                self.current_patch = self.current_slice[start_point_y : end_point_y, start_point_x : end_point_x]
                                # cv2.imwrite(f"s{i}r{j}c{k}.png", self.current_patch)
                                self.latest_analysis_results_dict = self.run_texture_analysis_algs(self.current_patch, np.ones_like(self.current_patch), analysis_setting_dict).copy()
                                self.latest_analysis_results_dict['patch_location'] = f"s{i}r{j}c{k}"
                                print(f"patch s{i}r{j}c{k} done")
                                try:
                                    # sheet_analysis_results_z_axis.append(list(self.latest_analysis_results_dict.values()) ) # slice i row j column k
                                    params_list = list(self.results_dict_template.keys())
                                    current_row = sheet_analysis_results_z_axis.max_row+1

                                    # sheet_analysis_results_z_axis.cell(row=current_row, column=1,value= f"s{i}r{j}c{k}")# slice i row j column k

                                    for param in self.latest_analysis_results_dict.keys():
                                        if param in params_list:
                                            sheet_analysis_results_z_axis.cell(row=current_row,
                                                                            column=params_list.index(param)+1,
                                                                            value= str(self.latest_analysis_results_dict[param]) )
                                    wb.save(result_excel_file)
                                    
                                except Exception as e:
                                    print('error in the results excel file:', e)
        
                    self.analysis_of_wholeScan_finished_sig.emit({"file_name":file_name, "scan_number_in_dataset":file_index, "total_scans":number_of_available_files})
                    file_index +=1

    def save_single_img_results(self, path):
        result_excel_file = path +'.xlsx'
        wb = Workbook()         
        wb.remove(wb.active)
        ws1 = wb.create_sheet('Info')
        ws1['A1'] = 'Scan File Name'
        ws1['A2'] = 'Mask File Name'
        sheet_analysis_results_z_axis = wb.create_sheet('Z_Axis_AnalysisResults')
        wb.save(result_excel_file)

    def check_number_of_scanfiles_in_path(self, path):
        file_index = 0
        for scan_folder in os.listdir(path):
            scan_folder_path = os.path.join(path, scan_folder)
            for file_name in os.listdir(scan_folder_path):
                if 'img' in file_name and (file_name.endswith(".nii.gz") or file_name.endswith(".nii")) :
                    file_index +=1
        return file_index
    

    def run(self):
        self.exec_()

        # while self._run_flag:
        #     QThread.msleep(1)


    def stop(self):
        """Sets run flag to False and waits for thread to finish"""
        self._run_flag = False
        self.wait()
    
    def create_results_dict_template(self, analysis_settings_dict):
        self.results_dict_template = dict()

        self.results_dict_template.update({'patch_location':'s0r0c0'})

        if 'GLCM' in analysis_settings_dict['available_texture_analysis_methods']:
            self.results_dict_template.update(glcm.RadiomicsGLCM.getFeatureNames())
        if 'FIRST_ORDER' in analysis_settings_dict['available_texture_analysis_methods']:
            self.results_dict_template.update(firstorder.RadiomicsFirstOrder.getFeatureNames())
        if 'GLRLM' in analysis_settings_dict['available_texture_analysis_methods']:
            self.results_dict_template.update(glrlm.RadiomicsGLRLM.getFeatureNames())
        if 'GLSZM' in analysis_settings_dict['available_texture_analysis_methods']:
            self.results_dict_template.update(glszm.RadiomicsGLSZM.getFeatureNames())
        if 'GLDM' in analysis_settings_dict['available_texture_analysis_methods']:
            self.results_dict_template.update(gldm.RadiomicsGLDM.getFeatureNames())

        print (self.results_dict_template)

    def runOn3dImage(self, data):
        num_slices = data.shape[0]
        for i in range(num_slices):
            self.runWholeAnalysis()

    def run_texture_analysis_algs(self, img_2d, mask_2d, settigs_dict): 
        results = dict()
        self.results_dict_template.update({'patch_location':'s0r0c0'})

        if 'GLCM' in settigs_dict['available_texture_analysis_methods']:
            results.update(self.runGLCM(img_2d, mask_2d, settigs_dict["radiomics_algs_settings"]))
        if 'FIRST_ORDER' in settigs_dict['available_texture_analysis_methods']:
            results.update(self.runFirstOrderFeatureCalcs(img_2d, mask_2d, settigs_dict["radiomics_algs_settings"]))
        if 'GLRLM' in settigs_dict['available_texture_analysis_methods']:
            results.update(self.runGLRLM(img_2d, mask_2d, settigs_dict["radiomics_algs_settings"]))
        if 'GLSZM' in settigs_dict['available_texture_analysis_methods']:
            results.update(self.runGLSZM(img_2d, mask_2d, settigs_dict["radiomics_algs_settings"]))
        if 'GLDM' in settigs_dict['available_texture_analysis_methods']:
            results.update(self.runGLDM(img_2d, mask_2d, settigs_dict["radiomics_algs_settings"]))
            
        return results
    

    def setImage(self, img_2d, mask_2d):
        img_2d = ((img_2d - img_2d.min()) * (1/(img_2d.max() - img_2d.min()) 
                                            * (2**16 -1) )).astype('uint16')

        mask_2d[mask_2d != 0] = 1000
        mask_2d = ((mask_2d - mask_2d.min()) * (1/(mask_2d.max() - mask_2d.min())
                                                    * (2**16 -1) )).astype('uint16')
        mask_2d[mask_2d != 0] = 1
        self.image = sitk.GetImageFromArray(img_2d)
        self.mask = sitk.GetImageFromArray(mask_2d)

    def setFeatureCalculationSettings(self):
        # testBinWidth = 25 this is the default bin size
        # testResampledPixelSpacing = [3,3,3] no resampling for now.
        applyLog = False
        applyWavelet = False

        # Setting for the feature calculation.
        # Currently, resampling is disabled.
        # Can be enabled by setting 'resampledPixelSpacing' to a list of 3 floats (new voxel size in mm for x, y and z)
        self.settings = {'binWidth': 25,
                    'interpolator': sitk.sitkBSpline,
                    'resampledPixelSpacing': None}

        #
        # If enabled, resample image (resampled image is automatically cropped.
        #
        interpolator = self.settings.get('interpolator')
        resampledPixelSpacing = self.settings.get('resampledPixelSpacing')
        if interpolator is not None and resampledPixelSpacing is not None:
            self.image, self.mask = imageoperations.resampleImage(self.image, self.mask, **self.settings)

        bb, correctedMask = imageoperations.checkMask(self.image, self.mask)
        if correctedMask is not None:
            self.mask = correctedMask
        self.image, self.mask = imageoperations.cropToTumorMask(self.image, self.mask, bb)

    def runFirstOrderFeatureCalcs(self, image, mask, settings):
        image = sitk.GetImageFromArray(image)
        mask = sitk.GetImageFromArray(mask)
        firstOrderFeatures = firstorder.RadiomicsFirstOrder(image, mask, **settings)
        firstOrderFeatures.enableAllFeatures()

        results = firstOrderFeatures.execute()
        return results

    def runGLCM(self, image, mask, settings):
        image = sitk.GetImageFromArray(image)
        mask = sitk.GetImageFromArray(mask)

        glcmFeatures = glcm.RadiomicsGLCM(image, mask, **settings)
        glcmFeatures.enableAllFeatures()

        # print('Will calculate the following GLCM features: ')
        # for f in glcmFeatures.enabledFeatures.keys():
        #     print('  ', f)
        #     print(getattr(glcmFeatures, 'get%sFeatureValue' % f).__doc__)

        # print('Calculating GLCM features...')
        results = glcmFeatures.execute()
        # print('done')

        # print('Calculated GLCM features: ')
        # for (key, val) in six.iteritems(results):
        #     print('  ', key, ':', val)

        return results
   
    def runGLRLM(self, image, mask, settings):
        image = sitk.GetImageFromArray(image)
        mask = sitk.GetImageFromArray(mask)
        glrlmFeatures = glrlm.RadiomicsGLRLM(image, mask, **settings)
        glrlmFeatures.enableAllFeatures()

        results = glrlmFeatures.execute()
        return results

    def runGLSZM(self, image, mask, settings):
        image = sitk.GetImageFromArray(image)
        mask = sitk.GetImageFromArray(mask)
        glszmFeatures = glszm.RadiomicsGLSZM(image, mask, **settings)
        glszmFeatures.enableAllFeatures()

        results = glszmFeatures.execute()
        return results

    def runGLDM(self, image, mask, settings):
        image = sitk.GetImageFromArray(image)
        mask = sitk.GetImageFromArray(mask)
        glszmFeatures = gldm.RadiomicsGLDM(image, mask, **settings)
        glszmFeatures.enableAllFeatures()

        results = glszmFeatures.execute()
        return results


# image = sitk.GetImageFromArray(np.ones((10,10)) )
# mask = sitk.GetImageFromArray(np.ones((10,10)) )
# settings = {'binWidth': 25,
#                     'interpolator': sitk.sitkBSpline,
#                     'resampledPixelSpacing': None}
# # glcm, glrlm, glszm, imageoperations, shape, gldm
# glcmFeatures = firstorder.RadiomicsFirstOrder(image, mask, **settings)
# glcmFeatures.enableAllFeatures()
# print('Will calculate the following GLRLM features: ')
# for f in glcmFeatures.enabledFeatures.keys():
#     print('  ', f)
#     print(getattr(glcmFeatures, 'get%sFeatureValue' % f).__doc__)


# # Show FirstOrder features, calculated on a LoG filtered image
# #
# if applyLog:
#   sigmaValues = numpy.arange(5., 0., -.5)[::1]
#   for logImage, imageTypeName, inputKwargs in imageoperations.getLoGImage(image, mask, sigma=sigmaValues):
#     logFirstorderFeatures = firstorder.RadiomicsFirstOrder(logImage, mask, **inputKwargs)
#     logFirstorderFeatures.enableAllFeatures()
#     results = logFirstorderFeatures.execute()
#     for (key, val) in six.iteritems(results):
#       laplacianFeatureName = '%s_%s' % (imageTypeName, key)
#       print('  ', laplacianFeatureName, ':', val)
# #
# # Show FirstOrder features, calculated on a wavelet filtered image
# #
# if applyWavelet:
#   for decompositionImage, decompositionName, inputKwargs in imageoperations.getWaveletImage(image, mask):
#     waveletFirstOrderFeaturs = firstorder.RadiomicsFirstOrder(decompositionImage, mask, **inputKwargs)
#     waveletFirstOrderFeaturs.enableAllFeatures()
#     results = waveletFirstOrderFeaturs.execute()
#     print('Calculated firstorder features with wavelet ', decompositionName)
#     for (key, val) in six.iteritems(results):
#       waveletFeatureName = '%s_%s' % (str(decompositionName), key)
#       print('  ', waveletFeatureName, ':', val)